#! python3
#
# Clive Ward-Cron
# December 11, 2019
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import sys
import shutil
import pprint
import re

print(os.getcwd())

# The first arg is the script name
args = sys.argv[1:]

fileName = ' '.join(args) + ".xlsx"

print(fileName)

filePath = os.path.join(os.getcwd(), fileName)

if os.path.exists(filePath):
    # Open the AGA excel doc
    wb = openpyxl.load_workbook(filePath)
    sheet = wb[wb.sheetnames[0]]

    data = []

    for row in range(2, sheet.max_row + 1):
        # Add data to dictionary
        # Change the key to str
        data.append(str(sheet['A' + str(row)].value))

    # pprint.pprint(data)

    newBook = Workbook()

    newFileName = 'RevisedDocV3.xlsx'

    ws1 = newBook.active
    ws1.title = "split addresses"

    ws1["A1"] = "Name"
    ws1["B1"] = "Address"
    ws1["C1"] = "City"
    ws1["D1"] = "State"

    addressRegex = re.compile(r'(\d+(\w|\W)+)()')
    stateRegex = re.compile(r'([Aa][LKSZRAEPlkszraep]|[Cc][AOTaot]|[Dd][ECec]|[Ff][LMlm]|[Gg][AUau]|[Hh][Ii]|[Ii][ADLNadln]|[Kk][SYsy]|[Ll][Aa]|[Mm][ADEHINOPSTadehinopst]|[Nn][CDEHJMVYcdehjmvy]|[Oo][HKRhkr]|[Pp][ARWarw]|[Rr][Ii]|[Ss][CDcd]|[Tt][NXnx]|[Uu][Tt]|[Vv][AITait]|[Ww][AIVYaivy])$')
    streetTypeRegex = re.compile(r'A(?:CCESS|LLEY|PPROACH|R(?:CADE|TERY)|VE(?:NUE)?)|B(?:A(?:NK|SIN|Y)|E(?:ACH|ND)|L(?:DG|VD)|O(?:ULEVARD|ARDWALK|WL)|R(?:ACE|AE|EAK|IDGE|O(?:ADWAY|OK|W))|UILDING|YPASS)|C(?:A(?:NAL|USEWAY)|ENTRE(?:WAY)?|HASE|IRC(?:LET?|U(?:IT|S))|L(?:OSE)?|O(?:MMON|NCOURSE|PSE|R(?:NER|SO)|UR(?:SE|T(?:YARD)?)|VE)|R(?:ES(?:CENT|T)?|IEF|OSS(?:ING)?)|U(?:LDESAC|RVE))|D(?:ALE|EVIATION|IP|OWNS|R(?:IVE(?:WAY)?)?)|E(?:ASEMENT|DGE|LBOW|N(?:D|TRANCE)|S(?:PLANADE|TATE)|X(?:P(?:(?:RESS)?WAY)|TENSION))|F(?:AIRWAY|IRETRAIL|O(?:LLOW|R(?:D|MATION))|R(?:(?:EEWAY|ONT(?:AGE)?)))|G(?:A(?:P|RDENS?|TE(?:S|WAY)?)|L(?:ADE|EN)|R(?:ANGE|EEN|O(?:UND|VET?)))|H(?:AVEN|E(?:ATH|IGHTS)|I(?:GHWAY|LL)|UB|WY)|I(?:NTER(?:CHANGE)?|SLAND)|JUNCTION|K(?:EY|NOLL)|L(?:A(?:NE(?:WAY)?)?|IN(?:E|K)|O(?:O(?:KOUT|P)|WER))|M(?:ALL|E(?:A(?:D|NDER)|WS)|OTORWAY)|NOOK|O(?:UTLOOK|VERPASS)|P(?:A(?:R(?:ADE|K(?:LANDS|WAY)?)|SS|TH(?:WAY)?)|DE|IER|L(?:A(?:CE|ZA))?|O(?:CKET|INT|RT)|RO(?:MENADE|PERTY)|URSUIT)?|QUA(?:D(?:RANT)?|YS?)|R(?:AMBLE|D|E(?:ACH|S(?:ERVE|T)|T(?:REAT|URN))|I(?:D(?:E|GE)|NG|S(?:E|ING))|O(?:AD(?:WAY)?|TARY|U(?:ND|TE)|W)|UN)|S(?:(?:ER(?:VICE)?WAY)|IDING|LOPE|PUR|QUARE|T(?:EPS|RAND|R(?:EET|IP))?|UBWAY)|T(?:ARN|CE|ERRACE|HRO(?:UGHWAY|WAY)|O(?:LLWAY|P|R)|RA(?:CK|IL)|URN)|UNDERPASS|V(?:AL(?:E|LEY)|I(?:EW|STA))|W(?:A(?:LK(?:WAY)?|Y)|HARF|YND)')

    for row in range(2, len(data)):
        address = addressRegex.search(data[row])
        state = stateRegex.search(data[row])

        # print(address)
        # print(data[row][0])
        if not address == None and not data[row][0].isdigit():
            pos = data[row].strip().find(address.group().strip())
            abrv = ''
            reverse = ''
            reverseArray = []
            name = data[row][:pos]
            ad = address.group().strip()
            city = ''
            st = ''
            if not state == None:
                st = state.group().strip()
                ad = ad[:ad.find(st)].strip()

            # reverseArray = ad.split(' ')
            # reverseArray.reverse()
            # reverse = ' '.join(reverseArray)
            # pprint.pprint(reverse)
            # streetType = streetTypeRegex.search(ad)

            # if not streetType == None:
            # abrv = streetType.group().strip()
            # city = ad[ad.find(abrv):].strip()

            ws1["A" + str(row)] = name
            ws1["B" + str(row)] = ad
            ws1["C" + str(row)] = city
            ws1["D" + str(row)] = st
            # ws1["E" + str(row)] = reverse
        else:
            ws1["A" + str(row)] = data[row]

    newBook.save(filename=newFileName)
