#! python3
#
# Clive Ward-Cron
# December 16, 2019
import openpyxl
from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
import os
import sys
# import shutil
# import pprint

# The first arg is the script name
args = sys.argv[1:]

# Append the xlsx extension name.
fileName = ' '.join(args) + ".xlsx"

# Determine the file path
filePath = os.path.join(os.getcwd(), fileName)

# Ensure that the file exists before we continue
if os.path.exists(filePath):
    # Open the Excel doc for processing.
    # and select the first sheet of the doc
    wb = openpyxl.load_workbook(filePath)
    sheet = wb[wb.sheetnames[0]]

    # Initialize an empty array
    data = []
    addresses = 0
    # maxColumns = 0

    # Starting with an offset of 1 to avoid the column letter,
    # go through each cell and create the addresses in a list
    for row in range(1, sheet.max_row + 1):
        # Concat each line of the address with a tab
        # Basing each addressee off of "TO THE PARENTS OF:"
        if str(sheet['A' + str(row)].value).strip() == 'TO THE PARENTS OF:':
            addresses += 1
            data.append(str(sheet['A' + str(row)].value).strip())
        else:
            data[addresses - 1] += "\t"
            data[addresses - 1] += str(sheet['A' + str(row)].value).strip()
            tabCount = data[addresses - 1].count("\t")
            # maxColumns = tabCount if tabCount > maxColumns else maxColumns

    # Create a new workbook to enter data into
    newName = ' '.join(args) + ' Revised.xlsx'
    newBook = Workbook()

    ws = newBook.active
    ws.title = 'Converted From Labels'

    ws["A1"] = "Parents of"
    ws["B1"] = "Name"
    ws["C1"] = "Address 1"
    ws["D1"] = "Address 2"
    ws["E1"] = "CSZ"

    # Initialize and loop through data adding it to workbook sheet
    x = 0
    addressee = []
    blanks = ['', '', '', '', '']
    a, b, c, d, e = blanks
    for row in range(0, len(data)):
        # split the larger string by the tabs
        addressee = data[x].split("\t")

        # split and assign each item for the current row
        # to an appropriate column
        if len(addressee) == 5:
            a, b, c, d, e = addressee
        else:
            a, b, c, e = addressee

        # assign data to columns
        ws['A'+str(row+2)] = a
        ws['B'+str(row+2)] = b
        ws['C'+str(row+2)] = c
        ws['D'+str(row+2)] = d
        ws['E'+str(row+2)] = e

        # reset column temp vars and increment counting var
        a, b, c, d, e = blanks
        x += 1

    # Save the New Workbook
    newBook.save(filename=newName)

    # print(addresses)
    # print(len(data))
    # print(maxColumns+1)
    # print(a, b, c, d, e)
