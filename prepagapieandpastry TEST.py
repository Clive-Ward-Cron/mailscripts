#! python3
#
# Clive Ward-Cron
# November 13, 2019
# Recreated February 17, 2020

import openpyxl
import os
import sys
import shutil
import pprint
from datetime import date

cwd = os.getcwd()

# print(cwd)

# The first arg is the script name
args = sys.argv[1:]

"""
Code below is from older version that required the data and the job numbers
to be input manually.

Currently working on only requiring AGA Spreadsheet File name and
then pulling the dates and job names from there.

"""

AGAfile = ' '.join(args) + ".xlsx"

path = os.path.join(cwd, AGAfile)

# print(path)

if os.path.exists(path):
    print('Path Exists In Current Folder')
    # print(path)
elif os.path.exists(path=os.path.join(cwd, "AGA DATABASE", AGAfile)):
    print('Path Exists in AGA DATABASE Folder')
    path = os.path.join(cwd, "AGA DATABASE", AGAfile)
    # print(path)
else:
    print('File Not Found In Current Folder Or "AGA DATABASE" Folder')
    sys.exit()

wb = openpyxl.load_workbook(path)
sheet = wb[wb.sheetnames[0]]

"""
Create an Array for the job #'s and an Array for the mail dates
that will use the job # to associate the date.

"""
jobNums = []
mailDates = []
jobCol = ""
dateCol = ""
missingFiles = []
dupeFolders = []
jobs = 0

# Find the job # field and Mail Date fields
for col in tuple(sheet.columns):
    fieldName = col[0].value.capitalize()
    # print(fieldName)
    if fieldName == "Job #".capitalize() or fieldName == "Job Number".capitalize():
        jobCol = col
    elif fieldName == "Mail Date".capitalize():
        dateCol = col

# Ensure that the fields were found
if jobCol == "" or dateCol == "":
    print("proper fields for job # and mail dates were not found.")
    sys.exit()

for row in range(1, len(jobCol)):

    # Ensure there are no None values
    if dateCol[row].value == None or jobCol[row].value == None:
        continue

    mailDate = str(dateCol[row].value.month) + \
        "-" + str(dateCol[row].value.day)
    job = str(jobCol[row].value)
    folderName = mailDate + ' ' + job
    csvName = None
    filePath = None

    # Ensure that the folder doesn't already exist before making one
    if not os.path.exists(folderName):
        os.makedirs(folderName)
    else:
        dupeFolders.append(folderName)

    # Retrieve the file name and add extension
    csvName = job + " Applied General Agency_File.csv"

    # Determine the file path based off workind directory
    filePath = os.path.join(cwd, csvName)

    # Ensure file exists and is a file
    if os.path.isfile(filePath):

        # Check that the folder exists and the file isn't already there
        if os.path.isdir(folderName) and not os.path.isfile(os.path.join(folderName, os.path.basename(filePath))):
            shutil.move(filePath, folderName)
            jobs += 1
        else:
            print()
            print("The file already exists in ", folderName)
    else:
        missingFiles.append(job + "\t" + csvName)
        continue

# Display errors
if len(dupeFolders) > 0:
    print("\nThe following folder(s) already exist in the current directory: ")
    for folder in dupeFolders:
        print(folder)
if len(missingFiles) > 0:
    print("\nThe following file(s) could not be found in the working directory: ")
    for job in missingFiles:
        print(job)
if jobs > 0:
    print("\nJob(s) successfully moved: " + str(jobs))

sys.exit()
# END TEST
################################################################################################################################################################
# Check that the AGA excel list exists.
if os.path.exists("\\\\wpnserver\\Data5\\374000-374499\\374243 - Acquire AGA\\Mailing\\AGA DATABASE\\Pie and Pastry.xlsx"):
    # Open the AGA excel doc
    wb = openpyxl.load_workbook(
        '\\\\wpnserver\\Data5\\374000-374499\\374243 - Acquire AGA\\Mailing\\AGA DATABASE\\Pie and Pastry.xlsx')

    # Create each needed sheet in a sheets array.
    # sheets = []
    # for num in sheetNums:
    #     sheets.append(wb[wb.sheetnames[num]])

    sheet = wb[wb.sheetnames[0]]

    # Create an array for the job #.
    data = []

    # Loop through and add data to the dict
    # for each sheet needed
    # Key is job# value is csv filename
    # for sheet in sheets:
    #     for row in range(2, sheet.max_row + 1):
    #         # Add data to dictionary
    #         # Change the key to str
    #         data[str(sheet['A' + str(row)].value)] = sheet['B' + str(row)].value

    for row in range(2, sheet.max_row + 1):
        # Add data to dictionary
        # Change the key to str
        data.append(str(sheet['J' + str(row)].value))

    # pprint.pprint(data)

    # Loop job#'s and concat the date to make folders
    # if the folders don't exist already.
    # Then search for the files in the current directory
    # and move the csv files to the correct folder.
    badJobs = []
    missingFiles = []
    dupeFolders = []
    jobs = 0
    for job in args:
        csvName = None
        filePath = None
        folderName = date + " " + job

        # Check if job# is in the data array
        if job in data:
            # Ensure that the folder doesn't already exist before making one
            if not os.path.exists(folderName):
                os.makedirs(folderName)
            else:
                dupeFolders.append(folderName)

            # Retrieve the file name and add extension
            csvName = job + " Applied General Agency_File.csv"

            # Determine the file path based off workind directory
            filePath = os.path.join(os.getcwd(), csvName)

            # Testing
            # print(csvName)
            # print(os.path.join(os.getcwd(), csvName))

            # Ensure file exists and is a file
            if os.path.isfile(filePath):
                # Testing
                # print()
                # print(filePath)

                # Check that the folder exists and the file isn't already there
                if os.path.isdir(folderName) and not os.path.isfile(os.path.join(folderName, os.path.basename(filePath))):
                    # Testing
                    # print("Move file...")
                    shutil.move(filePath, folderName)
                    # shutil.copy(filePath, folderName)
                    jobs += 1
                else:
                    print()
                    print("The file already exists in ", folderName)
            else:
                missingFiles.append(job + "\t" + csvName)
                continue
        else:
            badJobs.append(job)
            continue

    # Display errors
    if len(badJobs) > 0:
        print("\nThe following job(s) could not be found in the AGA job list: ")
        for job in badJobs:
            print(job)
    if len(dupeFolders) > 0:
        print("\nThe following folder(s) already exist in the current directory: ")
        for folder in dupeFolders:
            print(folder)
    if len(missingFiles) > 0:
        print("\nThe following file(s) could not be found in the working directory: ")
        for job in missingFiles:
            print(job)
    if jobs > 0:
        print("\nJob(s) successfully moved: " + str(jobs))
else:
    print("The AGA Excel Doc Does Not Exist")
