#! python3
# 
# Clive Ward-Cron
# November 13, 2019
import openpyxl
import os
import sys
import shutil
import pprint

# print(os.getcwd())

# The first arg is the script name
args = sys.argv[1:]

if len(args) <= 0:
    print("Please provide a date followed by job numbers")
    sys.exit()

# pop out the first value expected to be the date
date = args.pop(0)

if len(date) > 5 or len(date) < 3 or "-" not in date:
    print("Please provide a proper date ex. 11-28")
    sys.exit()

m, d = date.split("-")
if not m or not d:
    print("Please provide a proper date ex. 11-28")
    sys.exit()
elif int(m) not in range(1, 12+1) or int(d) not in range(1, 31+1):
    print("Please provide a proper date ex. 11-28")
    sys.exit()

if len(args) <= 0:
    print("please provide job numbers")
    sys.exit()


# Sort job numbers.
args.sort()

# pprint.pprint(args)
# print(date)

# Get a list of the sheets needed
sheetNums = []
for job in args:
    if int(job[0]) not in sheetNums:
        sheetNums.append(int(job[0]) - 1)
    else:
        continue

# Check that the AGA excel list exists.
if os.path.exists("\\\\wpnserver\\Data5\\367000-367499\\367050 Acquire AGA\\Mailing\\BigAgaRay\\367050 AGA (11-6-19).xlsx"):
    # Open the AGA excel doc
    wb = openpyxl.load_workbook('\\\\wpnserver\\Data5\\367000-367499\\367050 Acquire AGA\\Mailing\\BigAgaRay\\367050 AGA (11-6-19).xlsx')
    
    # Create each needed sheet in a sheets array.
    sheets = []
    for num in sheetNums:
        sheets.append(wb[wb.sheetnames[num]])
        
    # Create a dict for the job # and file name to be added to.
    data = {}

    # Loop through and add data to the dict
    # for each sheet needed
    # Key is job# value is csv filename
    for sheet in sheets:
        for row in range(2, sheet.max_row + 1):
            # Add data to dictionary
            # Change the key to str
            data[str(sheet['A' + str(row)].value)] = sheet['B' + str(row)].value

    # pprint.pprint(data)
    
    # Loop job#'s and concat the date to make folders
    # if the folders don't exist already.
    # Then search for the files in the current directory
    # and move the csv files to the correct folder.
    badJobs = []
    missingFiles = []
    dupeFolders = []
    for job in args:
        csvName = None
        filePath = None
        folderName = date + " " + job

        # Check if job# is in the data dict
        if job in data.keys():
            # Ensure that the folder doesn't already exist before making one
            if not os.path.exists(folderName):
                os.makedirs(folderName)
            else:
                dupeFolders.append(folderName)
                
            # Retrieve the file name and add extension
            csvName = data[job] + ".csv"
            
            # Determine the file path based off workind directory
            filePath = os.path.join(os.getcwd(), csvName)

            # Testing
            # print(csvName)
            # print(os.path.join(os.getcwd(), csvName))

            # Ensure file exists and is a file
            if os.path.isfile(filePath):
                #Testing
                #print()
                #print(filePath)

                # Check that the folder exists and the file isn't already there
                if os.path.isdir(folderName) and not os.path.isfile(os.path.join(folderName, os.path.basename(filePath))):
                    # Testing
                    # print("Move file...")
                    shutil.move(filePath, folderName)
                else:
                    print()
                    print("The file already exists in ", folderName)
            else:
                missingFiles.append(job + "\t" + csvName)
                continue
        else:
            badJobs.append(job)
            continue

    # Display errors
    if len(badJobs) > 0:
        print("\nThe following job(s) could not be found in the AGA job list: ")
        for job in badJobs:
            print(job)    
    if len(dupeFolders) > 0:
        print("\nThe following folder(s) already exist in the current directory: ")
        for folder in dupeFolders:
            print(folder)
    if len(missingFiles) > 0:
        print("\nThe following file(s) could not be found in the working directory: ")
        for job in missingFiles:
            print(job)
else:
    print("The AGA Excel Doc Does Not Exist")
